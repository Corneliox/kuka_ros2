import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE_ROOT = r"d:\~Ideas n Innovation\~~Taiwan\AU\11_Task_Robotic\kuka_ros2"
DEST_PATHS = [
    os.path.join(WORKSPACE_ROOT, "Journal", "Tabel_Referensi_Lengkap_17.xlsx"),
    r"D:\Download Move1\Jurnal_lolo\From Gilang\Revise\Tabel_Referensi_Lengkap_17.xlsx",
    r"D:\Download Move1\Jurnal_lolo\From Gilang\conf\Tabel_Referensi_Lengkap_17.xlsx"
]

def build_reference_workbook():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Sheet 1: Tabel 17 Referensi Terverifikasi
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Daftar 17 Referensi"
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = [
        "No",
        "Citation Key",
        "Penulis (Authors)",
        "Judul Lengkap Publikasi (Title)",
        "Jurnal / Konferensi / Penerbit",
        "Tahun",
        "DOI / URL Aktif"
    ]
    
    data1 = [
        [1, "ros2_2022", "S. Macenski, T. Foote, B. Gerkey, C. Lalancette, W. Woodall", "Robot Operating System 2: Design, architecture, and uses in the wild", "Science Robotics, Vol. 7, No. 66, eabm6074", 2022, "https://doi.org/10.1126/scirobotics.abm6074"],
        [2, "moveit2014", "D. Coleman, I. Sucan, S. Chitta, N. Correll", "Reducing the barrier to entry of complex robotic software: A MoveIt! case study", "Journal of Software Engineering for Robotics, Vol. 5, No. 1, pp. 3–16", 2014, "https://arxiv.org/abs/1404.3785"],
        [3, "jopenshowvar2015", "F. Sanfilippo, G. Marafioti, A. G. Robertson", "Controlling Kuka industrial robots: Flexible communication interface JOpenShowVar", "IEEE Robotics & Automation Magazine, Vol. 22, No. 4, pp. 96–109", 2015, "https://doi.org/10.1109/MRA.2015.2482839"],
        [4, "lbr_stack2024", "M. Huber et al.", "LBR-Stack: ROS 2 and Python integration of KUKA FRI for Med and IIWA robots", "Journal of Open Source Software (JOSS), Vol. 9, No. 103, p. 6138", 2024, "https://doi.org/10.21105/joss.06138"],
        [5, "villani2018", "V. Villani, F. Pini, F. Leali, C. Fantuzzi", "Survey on human–robot collaboration in industrial settings: Safety, intuitive interfaces and applications", "Mechatronics (Elsevier), Vol. 55, pp. 248–266", 2018, "https://doi.org/10.1016/j.mechatronics.2018.02.009"],
        [6, "trovato2020", "A. Rogowski et al.", "Integration of industrially-oriented human-robot speech communication and vision-based object recognition", "Sensors (MDPI), Vol. 20, No. 24, p. 7287", 2020, "https://doi.org/10.3390/s20247287"],
        [7, "cherubini2016", "A. Cherubini, R. Passama, A. Crosnier, A. Lasnier, P. Fraisse", "Collaborative manufacturing with physical human–robot interaction", "Robotics and Computer-Integrated Manufacturing, Vol. 40, pp. 1–14", 2016, "https://doi.org/10.1016/j.rcim.2015.12.007"],
        [8, "perzylo2016", "A. Perzylo, N. Somani, M. Rickert, A. Knoll", "Intuitive instruction of industrial robots: Semantic process descriptions for small lot production", "Proc. IEEE/RSJ International Conference on Intelligent Robots and Systems (IROS), pp. 3415–3422", 2016, "https://doi.org/10.1109/IROS.2016.7759358"],
        [9, "rt2_2023", "A. Brohan et al.", "RT-2: Vision-Language-Action models transfer web knowledge to robotic control", "Proc. 7th Conference on Robot Learning (CoRL), PMLR 229, pp. 2165–2183", 2023, "https://arxiv.org/abs/2307.15818"],
        [10, "openvla2024", "M. J. Kim et al.", "OpenVLA: An open-source vision-language-action model", "Proc. 8th Conference on Robot Learning (CoRL), PMLR 270, pp. 2679–2713", 2025, "https://arxiv.org/abs/2406.09246"],
        [11, "octo2024", "Octo Model Team et al.", "Octo: An open-source generalist robot policy", "Proc. Robotics: Science and Systems (RSS 2024)", 2024, "https://arxiv.org/abs/2405.12213"],
        [12, "kuka_kss2018", "KUKA Roboter GmbH", "KUKA System Software (KSS) 8.3 Operating and Programming Instructions for System Integrators", "KUKA AG Manual, Augsburg, Germany", 2018, "https://www.kuka.com"],
        [13, "kuka_eki2019", "KUKA Roboter GmbH", "KUKA.Ethernet KRL 3.1 Operating and Programming Instructions", "KUKA AG Manual, Augsburg, Germany", 2019, "https://www.kuka.com"],
        [14, "vosk2020", "N. Shmyrev", "Vosk: Offline speech recognition API", "Alpha Cephei Toolkit", 2020, "https://alphacephei.com/vosk"],
        [15, "aruco2014", "S. Garrido-Jurado, R. Muñoz-Salinas, F. J. Madrid-Cuevas, M. J. Marín-Jiménez", "Automatic generation and detection of highly reliable fiducial markers under occlusion", "Pattern Recognition (Elsevier), Vol. 47, No. 6, pp. 2280–2292", 2014, "https://doi.org/10.1016/j.patcog.2014.01.005"],
        [16, "gasparetto2012", "A. Gasparetto, P. Boscariol, A. Lanzutti, R. Vidoni", "Trajectory planning in robotics", "Mathematics in Computer Science (Springer), Vol. 6, No. 3, pp. 269–279", 2012, "https://doi.org/10.1007/s11786-012-0123-8"],
        [17, "yolo2016", "J. Redmon, S. Divvala, R. Girshick, A. Farhadi", "You Only Look Once: Unified, real-time object detection", "Proc. IEEE Conference on Computer Vision and Pattern Recognition (CVPR), pp. 779–788", 2016, "https://doi.org/10.1109/CVPR.2016.91"]
    ]
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    link_font = Font(name="Calibri", size=10.5, color="0563C1", underline="single")
    bold_key_font = Font(name="Consolas", size=10, bold=True, color="1F4E79")
    regular_font = Font(name="Calibri", size=10.5)
    italic_font = Font(name="Calibri", size=10.5, italic=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Write Headers
    ws1.append(headers1)
    ws1.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Write Data
    for row_idx, row_data in enumerate(data1, start=2):
        ws1.append(row_data)
        ws1.row_dimensions[row_idx].height = 24
        fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        # Col 1: No
        c1 = ws1.cell(row=row_idx, column=1)
        c1.alignment = center_align
        c1.font = Font(name="Calibri", size=10.5, bold=True)
        c1.fill = fill
        c1.border = thin_border
        
        # Col 2: Citation Key
        c2 = ws1.cell(row=row_idx, column=2)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.font = bold_key_font
        c2.fill = fill
        c2.border = thin_border
        
        # Col 3: Authors
        c3 = ws1.cell(row=row_idx, column=3)
        c3.alignment = left_align
        c3.font = regular_font
        c3.fill = fill
        c3.border = thin_border
        
        # Col 4: Title
        c4 = ws1.cell(row=row_idx, column=4)
        c4.alignment = left_align
        c4.font = italic_font
        c4.fill = fill
        c4.border = thin_border
        
        # Col 5: Venue
        c5 = ws1.cell(row=row_idx, column=5)
        c5.alignment = left_align
        c5.font = regular_font
        c5.fill = fill
        c5.border = thin_border
        
        # Col 6: Year
        c6 = ws1.cell(row=row_idx, column=6)
        c6.alignment = center_align
        c6.font = regular_font
        c6.fill = fill
        c6.border = thin_border
        
        # Col 7: URL / DOI with Hyperlink
        c7 = ws1.cell(row=row_idx, column=7)
        c7.hyperlink = row_data[6]
        c7.font = link_font
        c7.alignment = Alignment(horizontal="left", vertical="center")
        c7.fill = fill
        c7.border = thin_border

    # Set Column Widths for Sheet 1
    col_widths1 = {1: 6, 2: 18, 3: 32, 4: 42, 5: 36, 6: 8, 7: 38}
    for col_idx, width in col_widths1.items():
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # -------------------------------------------------------------
    # Sheet 2: Peta Penempatan Sitasi di Dokumen Word
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Peta Penempatan di Word")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = [
        "No Sitasi",
        "Bagian Dokumen (Section)",
        "Teks Asli di Word (Cari via Ctrl+F)",
        "Entri EndNote yang Dimasukkan / Aksi"
    ]
    
    data2 = [
        ["[7]", "1. Introduction (P8)", "rather than manual teach pendants Cherubini et al. (Collaborative Manufacturing).", "Ganti teks dalam kurung dengan: Cherubini et al. (2016)"],
        ["[6,7]", "1. Introduction (P8)", "reducing cognitive load on human scrub nurses Rogowski et al. (Speech-Vision HRI) & Cherubini.", "Ganti teks dalam kurung dengan: Rogowski et al. (2020) & Cherubini et al. (2016)"],
        ["[8]", "1. Introduction (P8)", "This capability keeps operator hands free for dexterous inspection tasks and significantly reduces workstation reconfiguration times.", "Sisipkan di akhir kalimat: Perzylo et al. (2016)"],
        ["[9]", "1. Introduction (P9)", "Recently, end-to-end Vision-Language-Action foundation models such as RT-2 Brohan et al. (RT-2),", "Ganti teks dalam kurung dengan: Brohan et al. (2023)"],
        ["[10]", "1. Introduction (P9)", "OpenVLA Kim et al. (OpenVLA),", "Ganti teks dalam kurung dengan: Kim et al. (2024)"],
        ["[11]", "1. Introduction (P9)", "and Octo Octo Model Team (Octo)", "Ganti teks dalam kurung dengan: Octo Model Team (2024)"],
        ["[3,5]", "1. Introduction (P10)", "rather than external sensor streams Sanfilippo (JOpenShowVar) & Villani (HRI Survey).", "Ganti teks dalam kurung dengan: Sanfilippo et al. (2015) & Villani et al. (2018)"],
        ["[1]", "1. Introduction (P11)", "governed by Robot Operating System 2 Macenski et al. (Science Robotics ROS 2).", "Ganti teks dalam kurung dengan: Macenski et al. (2022)"],
        ["[14]", "1. Introduction (P11)", "combines offline Vosk speech recognition Shmyrev (Vosk Offline ASR) with phonetic alias mapping,", "Ganti teks dalam kurung dengan: Shmyrev (2020)"],
        ["[15]", "1. Introduction (P11)", "perspective-corrected planar homography using a twelve-marker ArUco constellation Garrido-Jurado (ArUco Markers),", "Ganti teks dalam kurung dengan: Garrido-Jurado et al. (2014)"],
        ["[2]", "1. Introduction (P11) & Method (P62)", "and MoveIt 2 Coleman et al. (MoveIt!) / Robot motions are planned within MoveIt 2 [5]", "Ganti dengan: Coleman et al. (2014)"],
        ["[16]", "1. Introduction (P11) & Method (P62)", "with the Pilz Industrial Motion Planner Gasparetto et al. (Trajectory Planning) / Pilz Industrial Motion Planner [6]", "Ganti dengan: Gasparetto et al. (2012)"],
        ["[4]", "2. Method (P13)", "We were using KUKA KR6 R900-2 robot with ROS 2 Framework", "Sisipkan setelah 'ROS 2 Framework': Huber et al. (2024)"],
        ["[12]", "2. Method (P16)", "driven by a KUKA KRC4 controller running KSS 8.3.", "Sisipkan setelah 'KSS 8.3': KUKA Roboter GmbH (2018)"],
        ["[13]", "2. Method (P17)", "The ROS 2 host workstation communicates with the KRC4 controller over Gigabit Ethernet via IP address 192.168.1.147 on port 54600.", "Sisipkan setelah 'port 54600': KUKA Roboter GmbH (2019)"],
        ["[17]", "4. Discussion (P174)", "severe optical occlusions, or transparent reflective tools lacking distinct color tags.", "Sisipkan di akhir kalimat Limitations: Redmon et al. (2016)"]
    ]
    
    ws2.append(headers2)
    ws2.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx, row_data in enumerate(data2, start=2):
        ws2.append(row_data)
        ws2.row_dimensions[row_idx].height = 24
        fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        c1 = ws2.cell(row=row_idx, column=1)
        c1.alignment = center_align
        c1.font = Font(name="Consolas", size=10.5, bold=True, color="1F4E79")
        c1.fill = fill
        c1.border = thin_border
        
        c2 = ws2.cell(row=row_idx, column=2)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.font = Font(name="Calibri", size=10.5, bold=True)
        c2.fill = fill
        c2.border = thin_border
        
        c3 = ws2.cell(row=row_idx, column=3)
        c3.alignment = left_align
        c3.font = regular_font
        c3.fill = fill
        c3.border = thin_border
        
        c4 = ws2.cell(row=row_idx, column=4)
        c4.alignment = left_align
        c4.font = Font(name="Calibri", size=10.5, bold=True, color="004D40")
        c4.fill = fill
        c4.border = thin_border
        
    col_widths2 = {1: 12, 2: 30, 3: 65, 4: 48}
    for col_idx, width in col_widths2.items():
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Save across all destination directories
    for path in DEST_PATHS:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb.save(path)
        print(f"Successfully generated Excel file at: {path}")

if __name__ == "__main__":
    build_reference_workbook()
